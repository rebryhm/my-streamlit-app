import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import load_workbook

# Set page configuration
st.set_page_config(page_title="All Model Details", layout="wide")

# Function to load and preprocess the Excel file
@st.cache_data
def load_and_preprocess(file_path):
    # Load Excel with multi-row headers
    raw_data = pd.read_excel(file_path, header=[1, 2])  # Adjust to start from row 2 and capture multi-row headers

    # Flatten multi-level headers
    raw_data.columns = ['_'.join([str(x).strip() for x in col if pd.notna(x)]) for col in raw_data.columns]

    # Dynamically find the 'Customer' column
    customer_column = [col for col in raw_data.columns if "Customer_Unnamed: 1_level_1" in col]
    if not customer_column:
        raise KeyError("Customer column not found in the dataset.")
    
    # Fill merged cells in 'Customer' column
    raw_data[customer_column[0]] = raw_data[customer_column[0]].fillna(method='ffill')

    return raw_data, customer_column[0]  # Return both data and the 'Customer' column name

# File path
file_path = r"C:\Users\Nikita Rana\Desktop\Project\All Running Models Cell,Battery & PCM Spec Data Compilation.xlsx"

# Load and preprocess data
data, customer_column = load_and_preprocess(file_path)

# Streamlit App
st.title("All Model Details")
st.write("Compare models and add new data dynamically.")

# Display full dataset
if st.checkbox("Show Full Dataset"):
    st.write(data)

# Task 1: Add a New Model
st.header("Add a New Model")

# Create three columns for the layout
col1, col2, col3 = st.columns(3)

with st.form("add_model_form"):
    # Column 1: Customer and Model Information
    with col1:
        customer = st.text_input("Customer")
        model = st.text_input("Model")
        
        # Cell Dimensions
        cell_length = st.text_input("Cell Length (mm)")
        cell_width = st.text_input("Cell Width (mm)")
        cell_thickness = st.text_input("Cell Thickness (mm)")
        
    # Column 2: F1, F2, Tab-to-Tab Distance, and Total Cell Length
    with col2:
        f1 = st.text_input("F1 (mm)")
        f2 = st.text_input("F2 (mm)")
        tab_to_tab_distance = st.text_input("Tab-to-Tab Distance (mm)")
        total_cell_length = st.text_input("Total Length (Including Tab) (mm)")
        
        # Battery Dimensions
        battery_length = st.text_input("Battery Total Length (mm)")
        battery_width = st.text_input("Battery Width (mm)")
        
    # Column 3: Body and Head Thickness for Battery, PCM Dimensions
    with col3:
        body_thickness = st.text_input("Body Thickness (mm)")
        head_thickness = st.text_input("Head Thickness (mm)")
        
        pcm_length = st.text_input("PCM Total Length (mm)")
        pcm_board_length = st.text_input("PCM Board Length (mm)")
        fpc_length = st.text_input("FPC Length (mm)")
        pcm_width = st.text_input("PCM Width (mm)")
        
    # Submit Button
    submit_button = st.form_submit_button("Submit")

if submit_button:
    # Create the new row as a list (excluding the Customer column)
    new_row = [
        model,  # Model name
        cell_length, cell_width, cell_thickness, f1, f2, 
        tab_to_tab_distance, total_cell_length, battery_length, battery_width, 
        body_thickness, head_thickness, pcm_length, pcm_board_length, fpc_length, pcm_width
    ]

    try:
        # Load the workbook and select the active sheet
        wb = load_workbook(file_path)
        sheet = wb.active

        # Find the rows corresponding to the specified customer
        customer_rows = [i for i in range(2, sheet.max_row + 1) if sheet.cell(row=i, column=2).value == customer]

        if not customer_rows:
            st.warning(f"Customer '{customer}' not found in the Excel file. Appending the new model at the end.")
            # Append the new model data at the end
            sheet.append(new_row)
        else:
            # Insert the new row below the last row of the customer's models
            last_customer_row = customer_rows[-1] + 1
            sheet.insert_rows(last_customer_row)
            for col_num, value in enumerate(new_row, start=1):
                sheet.cell(row=last_customer_row, column=col_num, value=value)

        # Save the workbook
        wb.save(file_path)
        st.success("Model added successfully in the appropriate location!")
    except Exception as e:
        st.error(f"Error updating the Excel file: {e}")









# Task 2: Compare Two Models
st.header("Compare Two Models")
col1, col2 = st.columns(2)

# Select the first model and its customer
with col1:
    model1 = st.selectbox("Select First Model", options=data["Model_Unnamed: 0_level_1"].dropna().unique(), key="model1")
    model1_details = data[data["Model_Unnamed: 0_level_1"] == model1]
    customer1 = model1_details["Customer_Unnamed: 1_level_1"].iloc[0] if not model1_details.empty else "N/A"  # Fetch customer for Model 1

# Select the second model and its customer
with col2:
    model2 = st.selectbox("Select Second Model", options=data["Model_Unnamed: 0_level_1"].dropna().unique(), key="model2")
    model2_details = data[data["Model_Unnamed: 0_level_1"] == model2]
    customer2 = model2_details["Customer_Unnamed: 1_level_1"].iloc[0] if not model2_details.empty else "N/A"  # Fetch customer for Model 2

# Display the comparison details
st.subheader(f"Comparison of {model1} and {model2}")
col1, col2 = st.columns(2)

with col1:
    st.write(f"Details for {model1}:")
    st.write(model1_details)
    st.write(f"Customer: {customer1}")  # Show Customer for Model 1

with col2:
    st.write(f"Details for {model2}:")
    st.write(model2_details)
    st.write(f"Customer: {customer2}")  # Show Customer for Model 2

# Create a button to download the comparison sheet
if st.button("Get Comparison Sheet"):
    if model1_details.empty or model2_details.empty:
        st.error("One or both models do not have valid details. Please select valid models.")
    else:
        # Dynamically create the filename
        filename = f"{customer1}S{model1} Vs {customer2}S{model2}.xlsx"

        # Specify the folder for saving the file
        comparison_folder = r"C:\Users\Nikita Rana\Desktop\Project\Comparison sheets"
        
        # Ensure the folder exists
        import os
        os.makedirs(comparison_folder, exist_ok=True)

        # Full path to save the comparison sheet
        comparison_sheet_file = os.path.join(comparison_folder, filename)

        # Create the comparison sheet
        comparison_sheet = pd.concat([model1_details.T, model2_details.T], axis=1)
        comparison_sheet.columns = [model1, model2]

        # Save the comparison sheet to the specified folder
        comparison_sheet.to_excel(comparison_sheet_file, index=True)

        # Provide download button
        with open(comparison_sheet_file, "rb") as file:
            st.download_button(
                label=f"Download {filename}",
                data=file,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        st.success(f"Comparison sheet saved in 'Comparison sheets' folder: {filename}")
