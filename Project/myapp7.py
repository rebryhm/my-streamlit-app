# Import required libraries
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import os

# Set page configuration
st.set_page_config(page_title="All Model Details", layout="wide")

# Folder for file persistence
PERSISTENCE_FOLDER = "uploaded_files"
os.makedirs(PERSISTENCE_FOLDER, exist_ok=True)

# Function to load and preprocess Excel data
def load_and_preprocess(file_path):
    raw_data = pd.read_excel(file_path, header=1)
    return raw_data.fillna("").rename(columns=str.strip)

# Title of the app
st.title("All Model Details")
st.write("Add models dynamically and ensure they are inserted into the correct customer section.")

# Sidebar for file management
with st.sidebar:
    st.header("Excel File Management")

    # Upload Excel file
    uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])
    if uploaded_file:
        uploaded_file_path = os.path.join(PERSISTENCE_FOLDER, "uploaded_file.xlsx")
        with open(uploaded_file_path, "wb") as f:
            f.write(uploaded_file.getbuffer())
        st.success("File uploaded successfully!")
    else:
        uploaded_file_path = os.path.join(PERSISTENCE_FOLDER, "uploaded_file.xlsx")
        if not os.path.exists(uploaded_file_path):
            st.error("No file uploaded. Please upload the file first.")
        else:
            st.info("Using the previously uploaded file.")

# Form for adding a new model
st.header("Add a New Model")
col1, col2, col3 = st.columns(3)

with st.form("add_model_form"):
    # Input fields for the form
    with col1:
        customer = st.text_input("Customer")
        model = st.text_input("Model")
        cell_length = st.text_input("Cell Length (mm)")
        cell_width = st.text_input("Cell Width (mm)")
        cell_thickness = st.text_input("Cell Thickness (mm)")

    with col2:
        f1 = st.text_input("F1 (mm)")
        f2 = st.text_input("F2 (mm)")
        tab_to_tab_distance = st.text_input("Tab-to-Tab Distance (mm)")
        total_cell_length = st.text_input("Total Length (Including Tab) (mm)")
        battery_length = st.text_input("Battery Total Length (mm)")

    with col3:
        battery_width = st.text_input("Battery Width (mm)")
        body_thickness = st.text_input("Body Thickness (mm)")
        head_thickness = st.text_input("Head Thickness (mm)")
        pcm_length = st.text_input("PCM Total Length (mm)")
        pcm_board_length = st.text_input("PCM Board Length (mm)")
        fpc_length = st.text_input("FPC Length (mm)")
        pcm_width = st.text_input("PCM Width (mm)")

    submit_button = st.form_submit_button("Submit")

# Handle form submission
if submit_button and uploaded_file_path:
    new_row = [
        model, customer, cell_length, cell_width, cell_thickness, f1, f2, 
        tab_to_tab_distance, total_cell_length, battery_length, battery_width, 
        body_thickness, head_thickness, pcm_length, pcm_board_length, fpc_length, pcm_width
    ]

    try:
        # Load workbook and select the active sheet
        wb = load_workbook(uploaded_file_path)
        sheet = wb.active

        # Locate the last row for the specified customer
        last_customer_row = None
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if row[1].value == customer:  # Assuming 'Customer' is in the 2nd column
                last_customer_row = row[0].row

        # Determine where to insert the new row
        if last_customer_row:
            next_row = last_customer_row + 1  # Insert after the last customer's row
        else:
            next_row = sheet.max_row + 1  # Append at the end if customer not found

        # Insert new row at the determined position
        sheet.insert_rows(next_row)

        # Populate the new row with the model's details
        for col, value in enumerate(new_row, start=1):
            sheet.cell(row=next_row, column=col, value=value)

        # Save the updated workbook
        wb.save(uploaded_file_path)
        st.success("Model added successfully!")

        # Display the updated Excel data in the sidebar
        with st.sidebar:
            st.subheader("Updated Excel Data")
            updated_data = load_and_preprocess(uploaded_file_path)
            st.write(updated_data)

            # Provide a download option for the updated file
            with open(uploaded_file_path, "rb") as f:
                st.download_button(
                    label="Download Updated Excel File",
                    data=f,
                    file_name="Updated_Excel_File.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
    except Exception as e:
        st.error(f"Error updating the Excel file: {e}")