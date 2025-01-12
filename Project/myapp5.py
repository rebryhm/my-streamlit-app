import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import os
import tkinter as tk
from tkinter import filedialog

# Set page configuration
st.set_page_config(page_title="All Model Details", layout="wide")

# Create a folder for persistence
PERSISTENCE_FOLDER = "uploaded_files"
os.makedirs(PERSISTENCE_FOLDER, exist_ok=True)

# Function to load and preprocess the Excel file
#def load_and_preprocess(file_path):
#    raw_data = pd.read_excel(file_path)  # Load the Excel file
#   return raw_data.fillna("")  # Replace NaN with empty strings for clean display

def load_and_preprocess(file_path):
    # Specify the row containing the header explicitly
    raw_data = pd.read_excel(file_path, header=1)  # Adjust header row to 1 (second row in Excel)
    return raw_data.fillna("").rename(columns=str.strip)  # Replace NaN with empty strings and strip column names

# Title of the app
st.title("All Model Details")
st.write("Compare models and dynamically update dataset.")

# Sidebar for uploading and viewing Excel files
with st.sidebar:
    st.header("Excel File Management")
    
    # Step 1: Upload the Excel file (only once)
    uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])
    
    # Save the file to disk if uploaded
    if uploaded_file:
        uploaded_file_path = os.path.join(PERSISTENCE_FOLDER, "uploaded_file.xlsx")
        with open(uploaded_file_path, "wb") as f:
            f.write(uploaded_file.getbuffer())
        st.success("File uploaded successfully!")
    else:
        uploaded_file_path = os.path.join(PERSISTENCE_FOLDER, "uploaded_file.xlsx")
        if not os.path.exists(uploaded_file_path):
            st.error("No file uploaded. Please upload the file first.")
        else:
            st.info("File already uploaded and ready to be updated.")

# Main content: Add a new model
st.header("Add a New Model")
col1, col2, col3 = st.columns(3)

with st.form("add_model_form"):
    # Column 1: Customer and Model Information
    with col1:
        customer = st.text_input("Customer")
        model = st.text_input("Model")
        cell_length = st.text_input("Cell Length (mm)")
        cell_width = st.text_input("Cell Width (mm)")
        cell_thickness = st.text_input("Cell Thickness (mm)")
    
    # Column 2: F1, F2, Tab-to-Tab Distance, and Total Cell Length
    with col2:
        f1 = st.text_input("F1 (mm)")
        f2 = st.text_input("F2 (mm)")
        tab_to_tab_distance = st.text_input("Tab-to-Tab Distance (mm)")
        total_cell_length = st.text_input("Total Length (Including Tab) (mm)")
        battery_length = st.text_input("Battery Total Length (mm)")
        battery_width = st.text_input("Battery Width (mm)")
    
    # Column 3: Body and Head Thickness for Battery, PCM Dimensions
    with col3:
        body_thickness = st.text_input("Body Thickness (mm)")
        head_thickness = st.text_input("Head Thickness (mm)")
        pcm_length = st.text_input("PCM Total Length (mm)")
        pcm_board_length = st.text_input("PCM Board Length (mm)")
        fpc_length = st.text_input("FPC Length (mm)")
        pcm_width = st.text_input("PCM Width (mm)")

    # Submit Button
    submit_button = st.form_submit_button("Submit")

# Show the updated Excel data in the sidebar on the left side after the form is submitted
if submit_button and uploaded_file_path:
    # Correct the order of customer and model in the new row
    new_row = [
        model, customer, cell_length, cell_width, cell_thickness, f1, f2, 
        tab_to_tab_distance, total_cell_length, battery_length, battery_width, 
        body_thickness, head_thickness, pcm_length, pcm_board_length, fpc_length, pcm_width
    ]

    try:
        # Load the workbook and select the active sheet
        wb = load_workbook(uploaded_file_path)
        sheet = wb.active

        # Find the last row for the given customer
        last_customer_row = None
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if row[1].value == customer:  # Assuming customer is in the 2nd column
                last_customer_row = row[0].row  # Store the row number

        # Determine where to insert the new row
        if last_customer_row:
            next_row = last_customer_row + 1  # Insert after the last customer row
        else:
            next_row = sheet.max_row + 1  # If the customer doesn't exist, append at the end

        # Shift rows down to insert a new row at the specific position
        sheet.insert_rows(next_row)

        # Append data to the sheet at the specific position
        for col, value in enumerate(new_row, start=1):
            sheet.cell(row=next_row, column=col, value=value)

        # Save the updated workbook
        wb.save(uploaded_file_path)

        st.success("Model added successfully!")

        # Now, display the updated Excel data in the sidebar on the left side
        with st.sidebar:
            st.subheader("Updated Excel Data")
            updated_data = load_and_preprocess(uploaded_file_path)
            st.write(updated_data)

            # Allow download of the updated file with a unique key
            with open(uploaded_file_path, "rb") as f:
                st.download_button(
                    label="Download Updated Excel File",
                    data=f,
                    file_name="Updated_Excel_File.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_updated_file_button"  # Unique key for this button
                )
    except Exception as e:
        st.error(f"Error updating the Excel file: {e}")


st.header("Compare Two Models")

# Check if the updated file exists
if os.path.exists(uploaded_file_path):
    try:
        # Load the latest version of the uploaded/updated file
        latest_data = load_and_preprocess(uploaded_file_path)
        
    except Exception as e:
        st.error(f"Error processing the uploaded Excel file: {e}")

        # Display the dataset preview
        st.subheader("Latest Dataset Preview")
        st.write(latest_data)

        # Normalize column names for consistency
        latest_data.columns = latest_data.columns.str.strip().str.lower()
        model_column_name = "model"  # Expected column for Model
        customer_column_name = "customer"  # Expected column for Customer

        # Ensure required columns are present
        if model_column_name not in latest_data.columns or customer_column_name not in latest_data.columns:
            st.error(
                f"Required columns '{model_column_name}' and/or '{customer_column_name}' are missing. "
                "Please ensure your dataset includes these columns."
            )
        else:
            col1, col2 = st.columns(2)

            # Select the first model and its customer
            with col1:
                model1 = st.selectbox("Select First Model", options=latest_data[model_column_name].dropna().unique(), key="model1")
                model1_details = latest_data[latest_data[model_column_name] == model1]
                customer1 = model1_details[customer_column_name].iloc[0] if not model1_details.empty else "N/A"  # Fetch customer for Model 1

            # Select the second model and its customer
            with col2:
                model2 = st.selectbox("Select Second Model", options=latest_data[model_column_name].dropna().unique(), key="model2")
                model2_details = latest_data[latest_data[model_column_name] == model2]
                customer2 = model2_details[customer_column_name].iloc[0] if not model2_details.empty else "N/A"  # Fetch customer for Model 2

            # Display the comparison details
            st.subheader(f"Comparison of {model1} and {model2}")
            col1, col2 = st.columns(2)

            with col1:
                st.write(f"Details for {model1}:")
                st.write(model1_details)
                st.write(f"Customer: {customer1}")  # Show Customer for Model 1

            with col2:
                st.write(f"Details for {model2}:")
                st.write(model2_details)
                st.write(f"Customer: {customer2}")  # Show Customer for Model 2

import os
import pandas as pd
import streamlit as st

st.header("Compare Two Models")

# Check if the updated file exists
if os.path.exists(uploaded_file_path):
    try:
        # Load the latest version of the uploaded/updated file
        latest_data = load_and_preprocess(uploaded_file_path)

        # Display the dataset preview
        st.subheader("Latest Dataset Preview")
        st.write(latest_data)

        # Normalize column names for consistency
        latest_data.columns = latest_data.columns.str.strip().str.lower()
        model_column_name = "model"  # Expected column for Model
        customer_column_name = "customer"  # Expected column for Customer

        # Ensure required columns are present
        if model_column_name not in latest_data.columns or customer_column_name not in latest_data.columns:
            st.error(
                f"Required columns '{model_column_name}' and/or '{customer_column_name}' are missing. "
                "Please ensure your dataset includes these columns."
            )
        else:
            col1, col2 = st.columns(2)

            # Select the first model and its customer
            with col1:
                model1 = st.selectbox("Select First Model", options=latest_data[model_column_name].dropna().unique(), key="model1")
                model1_details = latest_data[latest_data[model_column_name] == model1]
                customer1 = model1_details[customer_column_name].iloc[0] if not model1_details.empty else "N/A"  # Fetch customer for Model 1

            # Select the second model and its customer
            with col2:
                model2 = st.selectbox("Select Second Model", options=latest_data[model_column_name].dropna().unique(), key="model2")
                model2_details = latest_data[latest_data[model_column_name] == model2]
                customer2 = model2_details[customer_column_name].iloc[0] if not model2_details.empty else "N/A"  # Fetch customer for Model 2

            # Display the comparison details
            st.subheader(f"Comparison of {model1} and {model2}")
            col1, col2 = st.columns(2)

            with col1:
                st.write(f"Details for {model1}:")
                st.write(model1_details)
                st.write(f"Customer: {customer1}")  # Show Customer for Model 1

            with col2:
                st.write(f"Details for {model2}:")
                st.write(model2_details)
                st.write(f"Customer: {customer2}")  # Show Customer for Model 2

            # Inside your "Get Comparison Sheet" button handler
            if st.button("Get Comparison Sheet"):
                if model1_details.empty or model2_details.empty:
                    st.error("One or both models do not have valid details. Please select valid models.")
                else:
                    try:
                        # Create the comparison sheet
                        comparison_sheet = pd.concat([model1_details.T, model2_details.T], axis=1)
                        comparison_sheet.columns = [model1, model2]

                        # Save the comparison sheet to a buffer
                        from io import BytesIO
                        buffer = BytesIO()
                        comparison_sheet.to_excel(buffer, index=True)
                        buffer.seek(0)

                        # Provide download button
                        st.download_button(
                            label="Download Comparison Sheet",
                            data=buffer,
                            file_name=f"{customer1}S{model1} Vs {customer2}S{model2}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                    except Exception as e:
                        st.error(f"Error generating the file: {e}")
    except Exception as e:
        st.error(f"Error processing the uploaded file for Task 2: {e}")
else:
    st.info("Please upload and update a dataset in Task 1 to proceed with Task 2.")
